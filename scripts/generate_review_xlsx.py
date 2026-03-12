"""
Generate a formatted Excel workbook for Sadie to review all 49 wizard
processes + 12 multi-intent scenarios.

Sheets:
  1) Single-Intent (49 rows) — quick scan with feedback column
  2) Multi-Intent (12 rows) — multi-tag scenarios
  3) Wizard Detail — one row per wizard step, linked by tag name

Outputs: review/parkm_wizard_review.xlsx
"""
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.services.classifier import EmailClassifier
from src.services.wizard import get_wizard_for_intent

# ── Styling constants ─────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
NOWRAP = Alignment(vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALT_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FEEDBACK_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_body(ws, num_cols, num_rows):
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL


# ── Same synthetic emails from generate_review_doc.py ─────────────────────
# (imported inline to keep this file self-contained)

SINGLE_INTENT_EMAILS = [
    {"target_tag": "Customer Canceling a Permit and Refunding", "subject": "Cancel my parking permit - need refund", "body": "Hi, I moved out of Riverside Apartments on March 1st 2026. My license plate is ABC1234. I need to cancel my parking permit and get a refund for the remaining balance. I was charged $75 on February 15th but I already moved out. Please help.", "from": "jane.doe@gmail.com"},
    {"target_tag": "Customer Inquiring for Grandfathered Permit", "subject": "Grandfathered permit question", "body": "Hello, I've been living at Oak Hill Community for over 3 years now. I was told my permit was grandfathered in under the old pricing. Can you confirm that my permit is still valid under the grandfathered rate? My plate is XYZ5678.", "from": "mike.smith@yahoo.com"},
    {"target_tag": "Customer Inquiring for Locked Down Permit", "subject": "Why is my permit locked?", "body": "I tried to renew my parking permit online but it says my permit is locked down. I live at Maple Grove unit 204. Can someone explain why it's locked and what I need to do? My plate is LMN9012.", "from": "sarah.jones@hotmail.com"},
    {"target_tag": "Customer Inquiring for Additional Permit", "subject": "Need a second parking permit", "body": "Hi there, I already have one permit for my Toyota but I just got a second car, a Honda Civic plate DEF3456. I live at Sunset Ridge apartment 112. Can I get an additional permit? What's the process and cost?", "from": "tom.williams@gmail.com"},
    {"target_tag": "Customer Rental Car", "subject": "Rental car temporary permit needed", "body": "My car is in the shop and I have a rental car for the next two weeks. The rental plate is RENT789. I live at Creekside Village unit 305. How do I get a temporary permit so I don't get towed? My regular plate is GHI7890.", "from": "lisa.brown@outlook.com"},
    {"target_tag": "Customer Double Charged or Extra Charges", "subject": "Charged twice for my permit this month!", "body": "I was charged $50 on March 1st and then another $50 on March 3rd for my parking permit at Willow Park. I should only have been charged once. My plate is JKL2345. Can you please look into this and refund the extra charge?", "from": "david.garcia@gmail.com"},
    {"target_tag": "Customer Guest Permit and Pricing Questions", "subject": "Guest parking permit info", "body": "Hi, I'm having family visit me at Pine Valley Apartments next weekend. How do I get a guest parking permit? How much does it cost and how long is it valid? I'm in unit 401.", "from": "emily.wilson@gmail.com"},
    {"target_tag": "Customer Miscellaneous Questions", "subject": "General question about parking", "body": "Hello, I just moved into Lakeside Terrace and I have a few questions about the parking situation. Is there assigned parking? Are there any visitor spots? What are the quiet hours for the parking garage? Thanks!", "from": "chris.martinez@yahoo.com"},
    {"target_tag": "Customer Sending Money Order", "subject": "Sending money order for parking permit", "body": "Hi, I don't have a credit card so I'd like to pay for my parking permit with a money order. I live at Brookfield Commons unit 210. Where do I mail the money order and who do I make it out to? The amount should be $45.", "from": "nancy.taylor@aol.com"},
    {"target_tag": "Customer Need help buying a permit", "subject": "How do I buy a parking permit?", "body": "I just signed my lease at Harbor Point Apartments and the leasing office said I need to buy a parking permit online. I went to the website but I'm confused about which permit to select. I'm in building C, unit 108. Can you walk me through the process?", "from": "kevin.anderson@gmail.com"},
    {"target_tag": "Customer Need Help Renewing Permit", "subject": "Permit renewal issue", "body": "My parking permit expired yesterday and I'm trying to renew it online but the system keeps giving me an error. I live at Elmwood Estates unit 502 and my plate is MNO4567. Can you help me renew my permit before I get towed?", "from": "amanda.thomas@gmail.com"},
    {"target_tag": "Customer Need Help Creating an Account", "subject": "Can't create my ParkM account", "body": "Hi, I'm a new resident at Stonegate Apartments unit 301. The leasing office told me to create an account on ParkM but when I try to register it says my email is already in use. I've never used ParkM before. My email is jessica.lee@gmail.com. Please help.", "from": "jessica.lee@gmail.com"},
    {"target_tag": "Customer New Towing Legislation No Parking", "subject": "New towing law question", "body": "I heard there's new legislation about towing in our area. I park at Ridgewood Community and I'm worried about how the new no-parking rules affect my vehicle. Can you explain what the new towing laws mean for residents? My plate is PQR6789.", "from": "ryan.clark@gmail.com"},
    {"target_tag": "Customer No Plate or Expired Tags", "subject": "My car has expired tags - will I get towed?", "body": "Hi, I'm waiting for my new registration to come in the mail. My current tags on my car are expired. I live at Cedar Springs unit 115 and my plate is STU8901. Will I get towed or ticketed while I wait for my new tags? What should I do?", "from": "michelle.lewis@yahoo.com"},
    {"target_tag": "Customer Someone is Parking in my Spot", "subject": "Someone keeps parking in my assigned spot!", "body": "There is a black SUV with plate VWX2345 that keeps parking in my assigned spot #42 at Fairview Apartments. This has happened three times this week. I've left notes but they keep doing it. What can be done about this? I'm in unit 208.", "from": "brian.walker@gmail.com"},
    {"target_tag": "Customer Parking Space Not in Dropdown", "subject": "My parking space isn't listed", "body": "I'm trying to register my vehicle for space #167 at Highland Park Apartments but the space number isn't showing up in the dropdown menu on the website. I'm in unit 403. Can you add my space to the system?", "from": "jennifer.hall@gmail.com"},
    {"target_tag": "Customer Password Reset", "subject": "Forgot my password", "body": "I can't remember my password for my ParkM account. I've tried the reset link but I'm not getting the email. My account email is daniel.young@gmail.com. Can you send me a password reset or reset it for me?", "from": "daniel.young@gmail.com"},
    {"target_tag": "Customer Towed Booted Ticketed", "subject": "MY CAR WAS TOWED!! HELP!!", "body": "I came outside this morning and my car is GONE. It was towed from the parking lot at Westwood Apartments. I have a valid permit! My plate is YZA3456 and I'm in unit 510. This is an emergency, I need my car for work. Who towed it and how do I get it back??", "from": "mark.king@gmail.com"},
    {"target_tag": "Customer Warned or Tagged", "subject": "Warning sticker on my car", "body": "I found a warning tag on my windshield at Valley View Apartments. It says my vehicle will be towed if not resolved in 48 hours. I have a valid permit for space #23. My plate is BCD4567, unit 102. Why did I get warned?", "from": "stephanie.wright@gmail.com"},
    {"target_tag": "Customer Payment Help", "subject": "Payment not going through", "body": "I'm trying to pay for my parking permit at Autumn Ridge but my credit card keeps getting declined. I've tried two different cards. I'm in unit 205. Is there something wrong with the payment system? Can I pay another way?", "from": "jason.green@yahoo.com"},
    {"target_tag": "Customer Update Vehicle Info", "subject": "Need to update my license plate", "body": "Hi, I just got a new car and need to update my license plate on my parking permit. Old plate: EFG5678, new plate: HIJ6789. I live at Magnolia Gardens unit 304. Can you update this for me?", "from": "laura.adams@gmail.com"},
    {"target_tag": "Customer Update Contact Info", "subject": "Update my phone number and email", "body": "Hi, I need to update my contact information on my ParkM account. My new phone number is 555-123-4567 and my new email is rachel.new@gmail.com. I live at Cypress Point unit 201. Thanks!", "from": "rachel.nelson@gmail.com"},
    {"target_tag": "Property Changing Resident Type for Approved Permit", "subject": "Change resident type for unit 305", "body": "Hi, this is the leasing office at Birchwood Estates. We need to change the resident type for the tenant in unit 305 from 'Standard' to 'Premium' for their approved permit. The resident is John Smith, plate KLM7890. Please update accordingly.", "from": "leasing@birchwoodestates.com"},
    {"target_tag": "Property Approving Grandfathered Permit", "subject": "Approve grandfathered permit - unit 412", "body": "Hello ParkM team, this is the property manager at Oakridge Manor. We'd like to approve a grandfathered permit for our long-term resident in unit 412, Maria Gonzalez. Her plate is NOP8901. She's been here since 2020 and qualifies for the old rate.", "from": "manager@oakridgemanor.com"},
    {"target_tag": "Property Approving Override Additional Permit", "subject": "Override approval for additional permit", "body": "Hi, this is Ashley from the front desk at Pinecrest Village. We're approving an override for an additional parking permit for unit 215, resident James Brown. His second vehicle plate is QRS9012. Please process this override.", "from": "frontdesk@pinecrestvillage.com"},
    {"target_tag": "Property Extending Expiration Date on a Permit", "subject": "Extend permit expiration for resident", "body": "Hello, this is the management office at Lakeshore Commons. We need to extend the permit expiration for resident in unit 608, Sarah Miller. Her current permit expires March 15th but her lease was extended through June 30th. Plate TUV1234.", "from": "office@lakeshorecommons.com"},
    {"target_tag": "Property Audits or Reports", "subject": "Monthly parking audit report needed", "body": "Hi ParkM, this is the property manager at Summit Hills. We need the monthly parking audit report for February 2026. Specifically, we need the number of active permits, expired permits, and any violations issued. Please send it as a spreadsheet if possible.", "from": "pm@summithills.com"},
    {"target_tag": "Property Checking if a Vehicle is Permitted", "subject": "Is this vehicle permitted?", "body": "Hi, this is maintenance at Greenfield Apartments. There's a red Ford F-150 with plate WXY2345 parked in lot B. Can you check if this vehicle has a valid permit? We want to verify before we call for a tow.", "from": "maintenance@greenfieldapts.com"},
    {"target_tag": "Property Checking Who Has a Space Number", "subject": "Who is assigned to space #55?", "body": "Hello, this is the leasing office at Riverdale Townhomes. We have a dispute about parking space #55. Can you tell us who is currently assigned to that space? We need to resolve this with our residents.", "from": "leasing@riverdaleth.com"},
    {"target_tag": "Property Checking Who is in a Unit", "subject": "Who is registered in unit 407?", "body": "Hi ParkM, this is the property manager at Courtyard Place. We need to know which residents are registered and have parking permits in unit 407. We're doing a lease audit and need to verify the information matches our records.", "from": "manager@courtyardplace.com"},
    {"target_tag": "Property Inquiring about a Tow Boot Ticket", "subject": "Tow inquiry for our property", "body": "Hello, this is management at Westgate Apartments. One of our residents is complaining that their car was towed from our lot last night. The vehicle is a blue Honda Civic plate ZAB3456. Can you provide details on why it was towed and which company performed the tow?", "from": "management@westgateapts.com"},
    {"target_tag": "Property Inquiring About a Warning Tag", "subject": "Warning tag placed on resident vehicle", "body": "Hi, this is the office at Meadowbrook Village. A resident in unit 103 received a warning tag on their vehicle, plate CDE4567. They're upset and came to us. Can you explain why the warning was issued and what steps the resident needs to take?", "from": "office@meadowbrookvillage.com"},
    {"target_tag": "Property Monitor Request", "subject": "Request parking lot monitoring", "body": "Hello ParkM, this is the HOA president at Eagle Ridge. We've been having issues with unauthorized vehicles parking overnight in our guest lot. Can we set up monitoring for lot C, especially between 10 PM and 6 AM? We'd like to start enforcement next week.", "from": "hoa@eagleridge.com"},
    {"target_tag": "Property Leasing Staff Login", "subject": "New leasing agent needs access", "body": "Hi, this is the property manager at Silver Creek Apartments. We have a new leasing agent, Brittany Cooper, who needs login access to the ParkM system. Her email is brittany.cooper@silvercreek.com. Can you set up her account?", "from": "pm@silvercreekapts.com"},
    {"target_tag": "Property Miscellaneous Questions", "subject": "General parking questions for our property", "body": "Hi ParkM, I'm the new property manager at Woodland Hills. I have a few general questions: How often do you patrol our lot? What's the process if we need to add more spaces? Do you provide signage? Can we get a copy of our current contract? Thanks.", "from": "newmanager@woodlandhills.com"},
    {"target_tag": "Property Sending Money Order", "subject": "Sending money order for resident permit", "body": "Hello, this is the office at Heritage Park. One of our residents in unit 502 doesn't have a bank account and would like to pay for their permit via money order. The amount is $60. Where should we send the money order and who do we make it payable to?", "from": "office@heritagepark.com"},
    {"target_tag": "Property Update or Register Employee Vehicles", "subject": "Register employee vehicles", "body": "Hi ParkM, this is HR at Cornerstone Communities. We need to register vehicles for three new maintenance employees: 1) John, plate FGH5678, 2) Maria, plate IJK6789, 3) Sam, plate LMN7890. They all need permits for lot A starting immediately.", "from": "hr@cornerstonecommunities.com"},
    {"target_tag": "Property Update Resident Vehicle", "subject": "Update resident vehicle info", "body": "Hi, leasing office at Springdale Apartments here. Our resident in unit 210, Michael Johnson, got a new car. Old plate: OPQ8901, new plate: RST9012. Can you update his parking permit with the new vehicle information?", "from": "leasing@springdalearpts.com"},
    {"target_tag": "Property Update Resident Contact Information", "subject": "Update resident contact info", "body": "Hello, this is the office at Foxwood Apartments. The resident in unit 318, Karen Davis, has a new phone number: 555-987-6543 and new email: karen.davis.new@gmail.com. Please update her ParkM account. Thanks!", "from": "office@foxwoodapts.com"},
    {"target_tag": "Property Update Resident Password", "subject": "Reset password for resident", "body": "Hi ParkM, this is the leasing office at Walnut Creek Apartments. Our resident in unit 105, Robert Wilson, has been locked out of his ParkM account. He says he's tried the forgot password link multiple times with no luck. Can you manually reset his password? His email is robert.wilson@email.com.", "from": "leasing@walnutcreekapts.com"},
    {"target_tag": "Property Register Resident Account for Them", "subject": "Create account for elderly resident", "body": "Hello, this is the office at Sunflower Senior Living. We have an elderly resident, Dorothy Thompson in unit 101, who is not tech-savvy and needs help creating her ParkM account. Her email is dorothy.t@gmail.com, phone 555-111-2222. Her vehicle is a silver Buick, plate UVW1234. Can you create her account for her?", "from": "office@sunflowersenior.com"},
    {"target_tag": "Property Cancel Resident Account", "subject": "Cancel resident account - moved out", "body": "Hi ParkM, this is the leasing office at Riverwalk Apartments. Resident in unit 404, Thomas Lee, moved out on March 5th. Please cancel his parking permit and account. His plate was XYZ2345. Thanks.", "from": "leasing@riverwalkapts.com"},
    {"target_tag": "Property Permitting PAID Resident Vehicle for Them", "subject": "Process paid permit for resident", "body": "Hi, the office at Crestview Condos here. We collected payment from our resident in unit 509, Angela Martinez, for a parking permit. Amount: $55. Her vehicle is a white Toyota Camry, plate ABC3456. Can you issue the permit on her behalf? She's already paid us directly.", "from": "office@crestviewcondos.com"},
    {"target_tag": "Property Resident Payment Help", "subject": "Resident having payment issues", "body": "Hello, this is the property manager at Bayview Apartments. Our resident in unit 602, Steven Chen, is having trouble making a payment for his parking permit through the website. He says his card keeps getting declined but it works everywhere else. His email is steven.chen@email.com. Can you assist?", "from": "manager@bayviewapts.com"},
    {"target_tag": "Property Guest Permits", "subject": "Guest permit policy for our community", "body": "Hi ParkM, this is the office at Palm Gardens. We're getting a lot of questions from residents about guest permits. Can you clarify: How many guest permits can each unit have? What's the cost? Is there a time limit? How do residents request them?", "from": "office@palmgardens.com"},
    {"target_tag": "Property Potential Leads", "subject": "Interested in ParkM for our community", "body": "Hello, I'm the HOA board president at Magnolia Square, a 200-unit apartment community. We're currently not using any parking management system and are interested in learning about ParkM's services. Can someone reach out to discuss pricing and setup? My direct number is 555-444-3333.", "from": "president@magnoliasquare.com"},
    {"target_tag": "Sales Rep Asking for a Vehicle to be Released", "subject": "Release vehicle from tow hold", "body": "Hey team, this is Jake from the towing division. We need to release the vehicle with plate DEF4567 that was flagged at Woodland Estates. The property manager confirmed it's a registered resident. Please release the hold so we can let it go.", "from": "jake@towingpartner.com"},
    {"target_tag": "Sales Rep Asking for a Vehicle to be Grandfathered", "subject": "Grandfather this vehicle in", "body": "Hi ParkM, this is Marcus from sales. The property at Hilltop Villas wants to grandfather in a vehicle for their long-term resident. Plate GHI5678, unit 201. The property confirmed they qualify under the old pricing. Can you process this?", "from": "marcus@parkmsales.com"},
    {"target_tag": "Towing or Monitoring Leads", "subject": "Parking enforcement services inquiry", "body": "Hi, I manage a commercial property complex in downtown. We're looking for a parking monitoring and towing enforcement partner. We have 3 buildings with about 500 parking spaces total. Would ParkM be able to provide monitoring and enforcement services? What's the process to get started?", "from": "info@downtownproperties.com"},
    {"target_tag": "The Law Asking for Information", "subject": "Law enforcement request for vehicle information", "body": "This is Officer Johnson, badge #4521, with the Metro Police Department. We are investigating a case (Case #2026-1234) and need information about a vehicle registered in your system. The plate is JKL6789. We need the registered owner's name and address associated with this permit. Please respond as soon as possible.", "from": "officer.johnson@metropd.gov"},
]

MULTI_INTENT_EMAILS = [
    {"label": "Cancel Permit + Refund + Password Reset", "subject": "Cancel permit, get refund, and can't log in", "body": "Hi, I moved out of Riverside Apartments on Feb 28th 2026. I need to cancel my parking permit and get a refund. Also, I can't log into my account to see my charges - I forgot my password. My plate is MNO1234 and I was charged $65 last month. Please help with all of this.", "from": "multi1@gmail.com"},
    {"label": "Double Charged + Cancel Permit", "subject": "Charged twice AND I'm moving out", "body": "I was charged $50 twice on March 1st for my permit at Willow Creek. That needs to be fixed. Also, I'm moving out on March 15th so I need to cancel my permit entirely. Plate: PQR2345, unit 303. This is really frustrating.", "from": "multi2@gmail.com"},
    {"label": "Update Vehicle + Renew Permit", "subject": "New car - need to update plate and renew", "body": "Hey, I just got a new car and my permit is about to expire. I need to update my plate from STU3456 to VWX4567 AND renew my parking permit at Summit Apartments unit 210. Can we do both at once?", "from": "multi3@gmail.com"},
    {"label": "Towed + Refund Request", "subject": "Car towed but I had a permit! I want a refund!", "body": "MY CAR WAS TOWED from Oakdale Apartments even though I have a valid permit!! Plate YZA5678, space #15, unit 404. I had to pay $250 to get my car back. I want a refund for the tow AND I want a refund on my parking permit because clearly it doesn't work. This is unacceptable!", "from": "multi4@gmail.com"},
    {"label": "Property: Check Vehicle + Check Unit", "subject": "Vehicle check and unit verification", "body": "Hi ParkM, this is the leasing office at Brookside Manor. Can you check two things: 1) Is the vehicle with plate BCD6789 permitted in our lot? And 2) Who is currently registered in unit 512? We're doing an end-of-month audit. Thanks!", "from": "leasing@brooksidemanor.com"},
    {"label": "Guest Permit + Someone in My Spot", "subject": "Guest permit and someone in my spot", "body": "Two issues: First, my parents are visiting this weekend and I need a guest permit for them at Meadow Lakes unit 207. Second, there's been a silver Honda with plate EFG7890 parking in my assigned spot #31 every day this week. Can you handle both?", "from": "multi6@gmail.com"},
    {"label": "Property: Register Account + Permit Payment", "subject": "New resident needs account and permit", "body": "Hello, this is the office at Cedar Point. We have a new move-in, Patricia Holmes, unit 603. She needs a ParkM account created (email: patricia.h@email.com, phone: 555-222-3333) AND we collected her permit payment of $45. Her vehicle is a blue Honda Accord plate HIJ8901. Please set up her account and issue the permit.", "from": "office@cedarpoint.com"},
    {"label": "Warning Tag + No Plate/Expired Tags", "subject": "Got a warning but my tags are being renewed", "body": "I got a warning tag on my car at Sunridge Apartments. The warning says my tags are expired, which they are, but I've already submitted my renewal to the DMV and I'm waiting for the new stickers. My plate is KLM9012, unit 108, space #22. What do I do so I don't get towed while I wait?", "from": "multi8@gmail.com"},
    {"label": "Property: Employee Vehicles + Leasing Staff Login", "subject": "New staff setup - vehicles and system access", "body": "Hi ParkM, this is HR at Grandview Properties. We have two new employees starting Monday. They both need: 1) Their vehicles registered for staff parking (John - plate NOP0123, Sarah - plate QRS1234), and 2) Login access to the ParkM management system. John's email: john@grandview.com, Sarah's email: sarah@grandview.com.", "from": "hr@grandviewproperties.com"},
    {"label": "Create Account + Buy Permit + Payment Help", "subject": "New here - can't figure anything out", "body": "Hi, I just moved into Lakeview Terrace unit 401 and I'm completely lost. I need to create a ParkM account, buy a parking permit, and I'm not sure what payment methods you accept. I tried to go to the website but I don't even know where to start. My car is a red Nissan plate TUV2345. Help!!", "from": "multi10@gmail.com"},
    {"label": "Property: Extend Permit + Change Resident Type", "subject": "Lease renewal - extend permit and change type", "body": "Hello ParkM, this is management at Ivy Court. Resident in unit 208, James White, renewed his lease and is upgrading from Standard to Premium parking. We need to: 1) Extend his permit expiration from March 31st to September 30th, and 2) Change his resident type to Premium. Plate WXY3456. Thanks!", "from": "management@ivycourt.com"},
    {"label": "Refund + Update Contact Info + Miscellaneous", "subject": "Moving out - refund, update info, and questions", "body": "Hi, I'm moving out of Parkside Heights on March 20th. A few things: 1) I need a refund on my remaining permit balance - I was charged $70 on March 1st. 2) My new email will be jane.newemail@gmail.com and phone 555-999-8888, please update that. 3) Also, do I need to return any parking stickers or key fobs? Plate ZAB4567, unit 506.", "from": "multi12@gmail.com"},
]


def format_entities_str(entities):
    """Format entities dict as a readable string."""
    if not entities:
        return ""
    parts = []
    for k, v in entities.items():
        if v:
            parts.append(f"{k}: {v}")
    return "; ".join(parts)


def summarize_wizard_steps(wizard):
    """One-line-per-step summary for the overview sheets."""
    lines = []
    for step in wizard.get("steps", []):
        sid = step.get("id", "?")
        text = step.get("text", "")
        dp = " [DECISION]" if step.get("decision_point") else ""
        req = " *" if step.get("required") else ""
        lines.append(f"{sid}. {text}{dp}{req}")
    return "\n".join(lines)


def get_templates_str(wizard):
    """List quick templates as a string."""
    templates = wizard.get("quick_templates", [])
    if not templates:
        return ""
    return "\n".join(f"- {t.get('label', '')} ({t.get('file', '')})" for t in templates)


def get_validation_str(wizard):
    """List validation items as a string."""
    items = wizard.get("validation_on_close", [])
    if not items:
        return ""
    return "\n".join(f"- {v}" for v in items)


def classify(classifier, subject, body, sender):
    result = classifier.classify_email(subject, body, sender)
    routing = classifier.get_routing_recommendation(result)
    return result, routing


def build_workbook(classifier):
    wb = Workbook()

    # ── Sheet 1: Single-Intent Overview ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "Single-Intent (49)"

    headers1 = [
        "#", "Target Tag", "Subject", "Email Body", "From",
        "Classified Tags", "Match?", "Confidence", "Complexity", "Urgency",
        "Key Entities", "Human Review", "Routing",
        "Wizard Steps (Summary)", "Response Templates",
        "Sadie's Feedback", "Status"
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    # Column widths
    col_widths_1 = [4, 35, 30, 60, 25, 35, 8, 10, 10, 8, 35, 12, 18, 50, 30, 30, 12]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    all_wizard_rows = []  # for Sheet 3

    for idx, email in enumerate(SINGLE_INTENT_EMAILS, 1):
        target = email["target_tag"]
        subject = email["subject"]
        body = email["body"]
        sender = email["from"]

        print(f"  [{idx}/49] {target}")
        classification, routing = classify(classifier, subject, body, sender)
        tags = classification.get("tags", [])
        confidence = classification.get("confidence", 0)
        if isinstance(confidence, float) and confidence <= 1.0:
            confidence = round(confidence * 100)
        entities = classification.get("key_entities", {})
        matched = target in tags

        # Get wizard for primary tag
        primary_wizard = get_wizard_for_intent(tags[0] if tags else target, classification)
        steps_summary = summarize_wizard_steps(primary_wizard)
        templates_str = get_templates_str(primary_wizard)

        row = idx + 1
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=target)
        ws1.cell(row=row, column=3, value=subject)
        ws1.cell(row=row, column=4, value=body)
        ws1.cell(row=row, column=5, value=sender)
        ws1.cell(row=row, column=6, value="; ".join(tags))
        match_cell = ws1.cell(row=row, column=7, value="YES" if matched else "NO")
        match_cell.fill = MATCH_FILL if matched else MISMATCH_FILL
        match_cell.font = Font(name="Calibri", bold=True, size=10)
        ws1.cell(row=row, column=8, value=f"{confidence}%")
        ws1.cell(row=row, column=9, value=classification.get("complexity", ""))
        ws1.cell(row=row, column=10, value=classification.get("urgency", ""))
        ws1.cell(row=row, column=11, value=format_entities_str(entities))
        ws1.cell(row=row, column=12, value="Yes" if classification.get("requires_human_review") else "No")
        ws1.cell(row=row, column=13, value=routing)
        ws1.cell(row=row, column=14, value=steps_summary)
        ws1.cell(row=row, column=15, value=templates_str)
        # Feedback columns - yellow highlight
        fb_cell = ws1.cell(row=row, column=16, value="")
        fb_cell.fill = FEEDBACK_FILL
        status_cell = ws1.cell(row=row, column=17, value="")
        status_cell.fill = FEEDBACK_FILL

        ws1.row_dimensions[row].height = 80

        # Collect wizard detail rows
        for tag in tags:
            wizard = get_wizard_for_intent(tag, classification)
            for step in wizard.get("steps", []):
                decision_options = ""
                if step.get("decision_point"):
                    opts = step.get("options", [])
                    decision_options = " | ".join(
                        f"{o.get('label', '')} -> {o.get('next_template', '')}" for o in opts
                    )
                missing_tmpl = ""
                ma = step.get("missing_action", {})
                if ma:
                    missing_tmpl = f"{ma.get('label', '')} ({ma.get('template', '')})"

                all_wizard_rows.append({
                    "test_num": idx,
                    "test_type": "Single",
                    "tag": tag,
                    "step_id": step.get("id", ""),
                    "step_text": step.get("text", ""),
                    "substep": step.get("substep", ""),
                    "required": "Yes" if step.get("required") else "",
                    "entity_field": step.get("entity_field", ""),
                    "entity_value": step.get("entity_value", ""),
                    "entity_found": "Yes" if step.get("entity_found") else ("No" if step.get("entity_field") else ""),
                    "decision_point": "Yes" if step.get("decision_point") else "",
                    "decision_options": decision_options,
                    "template": step.get("template", ""),
                    "missing_action": missing_tmpl,
                    "email_link": step.get("email_link", ""),
                    "show_for_action": step.get("show_for_action", ""),
                })

        time.sleep(0.5)

    style_header_row(ws1, len(headers1))
    style_body(ws1, len(headers1), len(SINGLE_INTENT_EMAILS))

    # Add data validation for Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Correct,Needs Fix,Needs Discussion,Skipped"',
        allow_blank=True,
    )
    dv.error = "Pick a status"
    dv.prompt = "Select review status"
    ws1.add_data_validation(dv)
    for r in range(2, len(SINGLE_INTENT_EMAILS) + 2):
        dv.add(ws1.cell(row=r, column=17))

    # ── Sheet 2: Multi-Intent Overview ────────────────────────────────────
    ws2 = wb.create_sheet("Multi-Intent (12)")

    headers2 = [
        "#", "Scenario", "Subject", "Email Body", "From",
        "Classified Tags", "Tag Count", "Confidence", "Complexity", "Urgency",
        "Key Entities", "Human Review", "Routing",
        "Wizard Steps (All Tags)", "Response Templates",
        "Sadie's Feedback", "Status"
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)

    col_widths_2 = [4, 30, 30, 60, 25, 45, 8, 10, 10, 8, 35, 12, 18, 55, 30, 30, 12]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for idx, email in enumerate(MULTI_INTENT_EMAILS, 1):
        label = email["label"]
        subject = email["subject"]
        body = email["body"]
        sender = email["from"]

        print(f"  [Multi {idx}/12] {label}")
        classification, routing = classify(classifier, subject, body, sender)
        tags = classification.get("tags", [])
        confidence = classification.get("confidence", 0)
        if isinstance(confidence, float) and confidence <= 1.0:
            confidence = round(confidence * 100)
        entities = classification.get("key_entities", {})

        # Combined wizard steps for all tags
        all_steps = []
        all_templates = []
        for tag in tags:
            wizard = get_wizard_for_intent(tag, classification)
            all_steps.append(f"--- {tag} ---")
            all_steps.append(summarize_wizard_steps(wizard))
            tmpl = get_templates_str(wizard)
            if tmpl:
                all_templates.append(f"[{tag}]\n{tmpl}")

            # Wizard detail rows
            for step in wizard.get("steps", []):
                decision_options = ""
                if step.get("decision_point"):
                    opts = step.get("options", [])
                    decision_options = " | ".join(
                        f"{o.get('label', '')} -> {o.get('next_template', '')}" for o in opts
                    )
                missing_tmpl = ""
                ma = step.get("missing_action", {})
                if ma:
                    missing_tmpl = f"{ma.get('label', '')} ({ma.get('template', '')})"

                all_wizard_rows.append({
                    "test_num": f"M{idx}",
                    "test_type": "Multi",
                    "tag": tag,
                    "step_id": step.get("id", ""),
                    "step_text": step.get("text", ""),
                    "substep": step.get("substep", ""),
                    "required": "Yes" if step.get("required") else "",
                    "entity_field": step.get("entity_field", ""),
                    "entity_value": step.get("entity_value", ""),
                    "entity_found": "Yes" if step.get("entity_found") else ("No" if step.get("entity_field") else ""),
                    "decision_point": "Yes" if step.get("decision_point") else "",
                    "decision_options": decision_options,
                    "template": step.get("template", ""),
                    "missing_action": missing_tmpl,
                    "email_link": step.get("email_link", ""),
                    "show_for_action": step.get("show_for_action", ""),
                })

        row = idx + 1
        ws2.cell(row=row, column=1, value=idx)
        ws2.cell(row=row, column=2, value=label)
        ws2.cell(row=row, column=3, value=subject)
        ws2.cell(row=row, column=4, value=body)
        ws2.cell(row=row, column=5, value=sender)
        ws2.cell(row=row, column=6, value="; ".join(tags))
        ws2.cell(row=row, column=7, value=len(tags))
        ws2.cell(row=row, column=8, value=f"{confidence}%")
        ws2.cell(row=row, column=9, value=classification.get("complexity", ""))
        ws2.cell(row=row, column=10, value=classification.get("urgency", ""))
        ws2.cell(row=row, column=11, value=format_entities_str(entities))
        ws2.cell(row=row, column=12, value="Yes" if classification.get("requires_human_review") else "No")
        ws2.cell(row=row, column=13, value=routing)
        ws2.cell(row=row, column=14, value="\n".join(all_steps))
        ws2.cell(row=row, column=15, value="\n".join(all_templates))
        fb_cell = ws2.cell(row=row, column=16, value="")
        fb_cell.fill = FEEDBACK_FILL
        status_cell = ws2.cell(row=row, column=17, value="")
        status_cell.fill = FEEDBACK_FILL

        ws2.row_dimensions[row].height = 120

        time.sleep(0.5)

    style_header_row(ws2, len(headers2))
    style_body(ws2, len(headers2), len(MULTI_INTENT_EMAILS))

    # Status dropdown for multi-intent sheet
    dv2 = DataValidation(
        type="list",
        formula1='"Correct,Needs Fix,Needs Discussion,Skipped"',
        allow_blank=True,
    )
    ws2.add_data_validation(dv2)
    for r in range(2, len(MULTI_INTENT_EMAILS) + 2):
        dv2.add(ws2.cell(row=r, column=17))

    # ── Sheet 3: Wizard Step Detail ───────────────────────────────────────
    ws3 = wb.create_sheet("Wizard Detail")

    headers3 = [
        "Test #", "Type", "Tag", "Step #", "Step Text", "Substep/Details",
        "Required?", "Entity Field", "Entity Value", "Entity Found?",
        "Decision Point?", "Decision Options", "Template",
        "Missing Action", "Email Link", "Show For Action",
        "Sadie's Feedback"
    ]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)

    col_widths_3 = [7, 7, 35, 6, 40, 50, 8, 14, 16, 10, 10, 40, 25, 25, 25, 15, 30]
    for i, w in enumerate(col_widths_3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    for i, wr in enumerate(all_wizard_rows, 1):
        row = i + 1
        ws3.cell(row=row, column=1, value=str(wr["test_num"]))
        ws3.cell(row=row, column=2, value=wr["test_type"])
        ws3.cell(row=row, column=3, value=wr["tag"])
        ws3.cell(row=row, column=4, value=str(wr["step_id"]))
        ws3.cell(row=row, column=5, value=wr["step_text"])
        ws3.cell(row=row, column=6, value=wr["substep"])
        ws3.cell(row=row, column=7, value=wr["required"])
        ws3.cell(row=row, column=8, value=wr["entity_field"])
        ws3.cell(row=row, column=9, value=wr["entity_value"] or "")
        ws3.cell(row=row, column=10, value=wr["entity_found"])
        ws3.cell(row=row, column=11, value=wr["decision_point"])
        ws3.cell(row=row, column=12, value=wr["decision_options"])
        ws3.cell(row=row, column=13, value=wr["template"])
        ws3.cell(row=row, column=14, value=wr["missing_action"])
        ws3.cell(row=row, column=15, value=wr["email_link"])
        ws3.cell(row=row, column=16, value=wr["show_for_action"])
        fb = ws3.cell(row=row, column=17, value="")
        fb.fill = FEEDBACK_FILL

    style_header_row(ws3, len(headers3))
    style_body(ws3, len(headers3), len(all_wizard_rows))

    print(f"\n  Wizard detail rows: {len(all_wizard_rows)}")
    return wb


def main():
    print("=" * 60)
    print("ParkM Wizard Review — Excel Generator")
    print("=" * 60)
    print()

    classifier = EmailClassifier()

    print("Classifying emails and building workbook...\n")
    wb = build_workbook(classifier)

    out_path = PROJECT_ROOT / "review" / "parkm_wizard_review.xlsx"
    wb.save(str(out_path))
    print(f"\nSaved: {out_path}")
    print(f"File size: {out_path.stat().st_size / 1024:.0f} KB")
    print("\nDone!")


if __name__ == "__main__":
    main()
